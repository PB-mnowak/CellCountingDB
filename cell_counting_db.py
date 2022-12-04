from os import path, listdir, getcwd, system, sep
from time import sleep
from copy import copy
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


def main():
    
    system('cls')
    db_path = get_db_path()

    if not check_db(db_path):
        print('---< Database file not found >---'.center(80))
        return None
    
    get_new_data(db_path)

def get_db_path() -> str:
    db_name = 'Cell counts - DB.xlsx'
    db_dir = path.dirname(path.abspath(__file__))
    return path.join(db_dir, db_name)

def check_db(db_path: str) -> bool:
    if path.exists(db_path):
        return True
    else:
        return False

def sheet_header(row) -> dict:
    header = [(col.value, i) for i, col in enumerate(row, 1) if col.value is not None]
    return dict(header)

def db_last_date(db_path: str) -> str:
    year = current_year()
    wb = load_workbook(db_path)
    sheet = wb[year]
    header = sheet_header(sheet[1])
    last_date = sheet.cell(row=2, column=header['Date']).value
    wb.close()
    return last_date

def load_date(date_str):
    date_format = '%d.%m.%Y %H:%M:%S'
    return datetime.strptime(date_str, date_format)

def current_year() -> str:
    now = datetime.now()
    date_format = '%Y'
    return datetime.strftime(now, date_format)

def extract_date(filename: str) -> float:
    year = filename[25:29]
    month = filename[23:25]
    day = filename[21:23]
    file_version = 0
    if '(' in filename:
        file_version = int(filename[31:].strip(').csv'))
    return float(f'{year}{month}{day}') + file_version / 100

def select_latest_csv(csv_path: str) -> str:
    csv_list = [file for file in listdir(csv_path) if file.endswith('.csv')]
    print(f'{len(csv_list)} CSV files detected')
    if csv_list:
        return max(csv_list, key=lambda x: extract_date(x))

def get_new_data(db_path: str):
    new_rows = 0
    csv_dir = getcwd()
    csv_file = select_latest_csv(csv_dir)
    if csv_file is not None:
        create_backup(db_path)
        csv_path = path.join(csv_dir, csv_file)
        last_date = db_last_date(db_path)
        df = import_data(csv_path, last_date)
        if df.shape[0] > 0:
            new_rows = update_db(db_path, df)
        transfer_csv(csv_path)
    update_status(new_rows)

def import_data(csv_path: str, last_date: str):
    with open(csv_path) as file:
        df_csv = pd.read_csv(file)
        df_csv['Date'] = pd.to_datetime(df_csv['Date'], format='%d/%m/%Y %H:%M')
        df_new = df_csv.loc[df_csv['Date'] > last_date]
    return df_new

def row_format():
    pass

def get_formatting(sheet, header: dict):
            row_white = sheet[2]
            row_gray = sheet[3]
            cell_formats = dict.fromkeys(header.keys(), dict())
            num_formats = dict.fromkeys(header.keys(), '')
            align_formats = dict.fromkeys(header.keys(), '')
            for col, i in header.items():
                cell_formats[col]['font'] = copy(row_white[i-1].font)
                cell_formats[col]['border'] = copy(row_white[i-1].border)
                cell_formats[col]['fill_white'] = copy(row_white[i-1].fill)
                cell_formats[col]['fill_gray'] = copy(row_gray[i-1].fill)
                num_formats[col] = copy(row_white[i-1].number_format)
                align_formats[col] = copy(row_white[i-1].alignment)
            return cell_formats, num_formats, align_formats

def fill_cell(sheet, i, header, cell_formats, num_formats, align_formats, df, new_rows):

            col_csv_dict = {
                'Name': 'Name',
                'Date': 'Date',
                'Total Cell\n[mvc/mL]': 'Total Cell',
                'Live Cell\n[mvc/mL]': 'Live Cell',
                'Dead Cell\n[mvc/mL]': 'Dead Cell',
                'Viability': 'Viability',
                'Average Cell Size': 'Average Cell Size',
                'Total Num': 'Total Num',
                'Live Num': 'Live Num',
                'Dead Num': 'Dead Num',
                'Protocol': 'Protocol',
                }
            
            n = i + 2
            for col in header.keys():
                db_cell =  sheet.cell(n, header[col])
                # print([a for a in dir(db_cell) if not a.startswith('_')])
                db_cell.font = cell_formats[col]['font']
                db_cell.border = cell_formats[col]['border']

                if (n + new_rows) % 2 == 0:
                    db_cell.fill = cell_formats[col]['fill_white']
                else:
                    db_cell.fill = cell_formats[col]['fill_gray']
                
                csv_col = col_csv_dict.get(col)
                if csv_col is not None:
                    csv_value = df.at[i, csv_col]
                    if 'mvc/mL' in col:
                        csv_value /= 10**6
                    elif 'Viability' in col:
                        csv_value = float(csv_value.strip('%')) / 100
                    db_cell.value = csv_value
                db_cell.number_format = num_formats[col]
                db_cell.alignment = align_formats[col]
                db_cell.hyperlink = None

def update_status(new_rows: int):
    print(f'{new_rows} new records added')

def update_db(db_path: str, df):
    year = current_year()
    wb = load_workbook(db_path)
    sheet = wb[year]
    header = sheet_header(sheet[1])
    cell_formats, num_formats, align_formats = get_formatting(sheet, header)
    new_rows = df.shape[0]
    if new_rows > 0:
        sheet.insert_rows(2, new_rows)
        for i in range(new_rows):
            fill_cell(sheet, i, header, cell_formats, num_formats, align_formats, df, new_rows)
    save_db(wb, db_path)
    wb.close()
    return new_rows

def save_db(wb, db_path: str):
    try:
        wb.save(db_path)
    except IOError:
        print('---> Unable to save. File was opened by other user.')
        while True:
            try_save = input('Try to save later? (Y/N) ').upper()
            if try_save == 'Y':
                print('Leave the window open. Attempts to save will be made every 60 sec.')
                save_later(wb, db_path)
                break
            elif try_save == 'N':
                break

def save_later(wb, db_path: str):
    while True:
        sleep(60)
        try:
            wb.save(db_path)
            break
        except:
            pass

def transfer_csv(csv_path: str):
    csv_folder = r'P:\_research group folders\PT Proteins\PT4\_PT4_cell_counting\CSV files'
    if path.exists(csv_folder):
        system(f'move "{csv_path}" "{csv_folder}\\" >nul')
        print(f'{csv_path.split(sep)[-1]} transferred to PB_all')

def create_backup(db_path: str):
    today = datetime.today().strftime('%Y%m%d_')
    db_dir, db_name = path.split(db_path)

    backup_n = 5
    backup_dir = path.join(db_dir, 'Backup')
    backup_new = path.join(backup_dir, today + db_name)
    system(f'copy "{db_path}" "{backup_new}" > nul')
    remove_old_backup(backup_dir, backup_n)

def remove_old_backup(backup_dir: str, backup_num: int):
    if len(backup_list := listdir(backup_dir)) > backup_num:
        try:
            oldest_file = min(backup_list, key=lambda x: int(x.split('_')[0]))
            system(f'del "{path.join(backup_dir, oldest_file)}"')
        except Exception as e:
            print(e)


if __name__ == '__main__':
    main()
