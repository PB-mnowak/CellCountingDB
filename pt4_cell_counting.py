import pandas as pd
from copy import copy
from os import listdir, getcwd, system, sep
from os.path import exists, join
from datetime import datetime
from openpyxl import load_workbook


def main():
    
    system('cls')
    new_rows = 0
    
    db_path = r'P:\_research group folders\PT Proteins\PT4\_PT4_cell_counting'
    db_name = 'Cell counts - DB.xlsx'
    db_full_path = join(db_path, db_name)
    if not db_check(db_full_path):
        print('---< Database file not found >---'.center(80))
        return None
    
    file_path = getcwd()
    filename = select_latest_csv(file_path)
    if filename is not None:
        file_full_path = join(file_path, filename)
        last_date = db_last_date(db_full_path)
        df = import_data(file_full_path, last_date)
        if df.shape[0] > 0:
            new_rows = update_db(db_full_path, df)
        transfer_csv(file_full_path)
            
    update_status(new_rows)


def db_check(db_full_path):
    if exists(db_full_path):
        return True
    else:
        return False

def sheet_header(row):
    header = [(col.value, i) for i, col in enumerate(row, 1) if col.value is not None]
    return dict(header)

def db_last_date(db_full_path):
    year = this_year()
    wb = load_workbook(db_full_path)
    sheet = wb[year]
    header = sheet_header(sheet[1])
    last_date = sheet.cell(row=2, column=header['Date']).value
    wb.close()
    return last_date

def load_date(date_str):
    date_format = '%d.%m.%Y %H:%M:%S'
    return datetime.strptime(date_str, date_format)

def this_year():
    now = datetime.now()
    date_format = '%Y'
    return datetime.strftime(now, date_format)

def extract_date(filename):
    file_date = filename.strip('.csv')
    file_date = file_date.split('_')[-1]
    file_date = file_date.replace(' (', '.')
    file_date = file_date.strip(')')
    return float(file_date)

def select_latest_csv(file_path):
    csv_list = [file for file in listdir(file_path) if file.endswith('.csv')]
    if csv_list:
        return max(csv_list, key=lambda x: extract_date(x))

def import_data(file_full_path, last_date):
    with open(file_full_path) as file:
        df_csv = pd.read_csv(file)
        df_csv['Date']= pd.to_datetime(df_csv['Date'], format='%d/%m/%Y %H:%M')
        df_new = df_csv.loc[df_csv['Date'] > last_date]
    return df_new

def row_format():
    pass

def get_formatting(sheet, header):
            row_white = sheet[2]
            row_gray = sheet[3]
            cell_formats = dict.fromkeys(header.keys(), dict())
            num_formats = dict.fromkeys(header.keys(), '')
            align_formats = dict.fromkeys(header.keys(), '')
            for col, i in header.items():
                cell_formats[col]['font'] = copy(row_white[i-1].font)
                cell_formats[col]['border'] = copy(row_white[i-1].border)
                cell_formats[col]['fill_white'] = copy(row_white[i-1].fill)
                cell_formats[col]['fill_gray'] = copy(row_gray[i-1].fill)
                num_formats[col] = copy(row_white[i-1].number_format)
                align_formats[col] = copy(row_white[i-1].alignment)
            return cell_formats, num_formats, align_formats

def fill_cell(sheet, i, header, cell_formats, num_formats, align_formats, df, new_rows):

            col_csv_dict = {
                'Name': 'Name',
                'Date': 'Date',
                'Total Cell\n[mvc/mL]': 'Total Cell',
                'Live Cell\n[mvc/mL]': 'Live Cell',
                'Dead Cell\n[mvc/mL]': 'Dead Cell',
                'Viability': 'Viability',
                'Average Cell Size': 'Average Cell Size',
                'Total Num': 'Total Num',
                'Live Num': 'Live Num',
                'Dead Num': 'Dead Num',
                'Protocol': 'Protocol',
            }
            
            n = i + 2
            for col in header.keys():
                db_cell =  sheet.cell(n, header[col])
                # print([a for a in dir(db_cell) if not a.startswith('_')])
                db_cell.font = cell_formats[col]['font']
                db_cell.border = cell_formats[col]['border']

                if (n + new_rows) % 2 == 0:
                    db_cell.fill = cell_formats[col]['fill_white']
                else:
                    db_cell.fill = cell_formats[col]['fill_gray']
                
                csv_col = col_csv_dict.get(col, False)
                if csv_col is not False:
                    csv_value = df.at[i, csv_col]
                    if 'mvc/mL' in col:
                        csv_value /= 10**6
                    elif 'Viability' in col:
                        csv_value = float(csv_value.strip('%')) / 100
                    db_cell.value = csv_value
                db_cell.number_format = num_formats[col]
                db_cell.alignment = align_formats[col]
                db_cell.hyperlink = None

def update_status(new_rows):
    print(f'{new_rows} new records added')
    # print()

def update_db(db_full_path, df):
    year = this_year()
    wb = load_workbook(db_full_path)
    sheet = wb[year]
    header = sheet_header(sheet[1])
    cell_formats, num_formats, align_formats = get_formatting(sheet, header)
    new_rows = df.shape[0]
    if new_rows > 0:
        sheet.insert_rows(2, new_rows)
        for i in range(new_rows):
            fill_cell(sheet, i, header, cell_formats, num_formats, align_formats, df, new_rows)

    wb.save(db_full_path)
    wb.close()
    return new_rows

def transfer_csv(file_full_path):
    csv_folder = r'P:\_research group folders\PT Proteins\PT4\_PT4_cell_counting\CSV files'
    if exists(csv_folder):
        system(f'move {file_full_path} "{csv_folder}\\" >nul')
        print(f'{file_full_path.split(sep)[-1]} transferred to PB_all')

if __name__ == '__main__':
    main()
